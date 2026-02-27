# members/views_import.py
"""
Bulk Import Members from Excel
"""

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.core.exceptions import ValidationError
from .models import Member, Family, Sector
import openpyxl
from datetime import datetime
import datetime as dt
import re


@login_required
def bulk_import_view(request):
    """
    Upload Excel file untuk bulk import jemaat.
    
    Features:
    - Upload .xlsx file
    - Validate data
    - Preview before import
    - Auto-create Family & Sector
    - Row-by-row error reporting
    """
    
    if request.method == 'POST':
        if 'preview' in request.POST:
            # Step 1: Upload & Preview
            return handle_upload_and_preview(request)
        elif 'confirm_import' in request.POST:
            # Step 2: Confirm & Import
            return handle_confirm_import(request)
    
    context = {
        'step': 'upload',
    }
    return render(request, 'members/bulk_import.html', context)


def handle_upload_and_preview(request):
    """Handle file upload and show preview"""
    
    excel_file = request.FILES.get('excel_file')
    
    if not excel_file:
        messages.error(request, 'File Excel wajib di-upload.')
        return redirect('members:bulk_import')
    
    # Validate file extension
    if not excel_file.name.endswith('.xlsx'):
        messages.error(request, 'File harus berformat .xlsx')
        return redirect('members:bulk_import')
    
    try:
        # Parse Excel
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        
        # Parse rows
        parsed_data = []
        errors = []
        
        # Skip header row
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # Skip empty rows
                continue
            
            try:
                member_data = parse_row(row, row_num)
                # Convert dates to strings for JSON serialization
                member_data_serializable = {
                    k: v.isoformat() if isinstance(v, dt.date) else v
                    for k, v in member_data.items()
                }
                parsed_data.append(member_data_serializable)
            except ValidationError as e:
                errors.append({
                    'row': row_num,
                    'errors': e.message_dict if hasattr(e, 'message_dict') else str(e)
                })
        
        # Store in session for confirmation step
        request.session['import_data'] = parsed_data
        request.session['import_errors'] = errors
        
        context = {
            'step': 'preview',
            'data': parsed_data[:10],  # Show first 10 for preview
            'total_rows': len(parsed_data),
            'errors': errors,
            'has_errors': len(errors) > 0,
        }
        
        return render(request, 'members/bulk_import.html', context)
        
    except Exception as e:
        messages.error(request, f'Error membaca file Excel: {str(e)}')
        return redirect('members:bulk_import')


def handle_confirm_import(request):
    """Confirm and execute import"""
    
    import_data = request.session.get('import_data', [])
    
    if not import_data:
        messages.error(request, 'Tidak ada data untuk di-import.')
        return redirect('members:bulk_import')
    
    success_count = 0
    error_count = 0
    errors = []
    
    # Helper function to convert string dates back to date objects
    def parse_date_string(date_str):
        if not date_str:
            return None
        try:
            return datetime.strptime(date_str, '%Y-%m-%d').date()
        except:
            return None
    
    try:
        with transaction.atomic():
            for row_num, data in enumerate(import_data, start=2):
                try:
                    # Convert date strings back to date objects
                    tanggal_lahir = parse_date_string(data.get('tanggal_lahir'))
                    tanggal_baptis = parse_date_string(data.get('tanggal_baptis'))
                    tanggal_sidi = parse_date_string(data.get('tanggal_sidi'))
                    tanggal_nikah = parse_date_string(data.get('tanggal_nikah'))
                    tanggal_meninggal = parse_date_string(data.get('tanggal_meninggal'))
                    
                    # Get or create Sector
                    sector, _ = Sector.objects.get_or_create(
                        name=data['nama_sektor'],
                        defaults={'description': f'Auto-created from import'}
                    )
                    
                    # Get or create Family
                    family, family_created = Family.objects.get_or_create(
                        family_name=data['nama_keluarga'],
                        sector=sector,
                        defaults={
                            'address_street': data['alamat_jalan'],
                            'address_city': data['kota'],
                            'address_province': data['provinsi'],
                            'address_postal_code': data.get('kode_pos', ''),
                            'phone_number': data['telepon_keluarga'],
                        }
                    )
                    
                    # Check if member already exists (by name + family)
                    existing = Member.objects.filter(
                        full_name=data['nama_lengkap'],
                        family=family
                    ).first()
                    
                    if existing:
                        errors.append({
                            'row': row_num,
                            'error': f'Jemaat {data["nama_lengkap"]} sudah ada di keluarga ini'
                        })
                        error_count += 1
                        continue
                    
                    # Create Member
                    member = Member.objects.create(
                        full_name=data['nama_lengkap'],
                        gender=data['jenis_kelamin'],
                        date_of_birth=tanggal_lahir,
                        family=family,
                        current_sector=sector,
                        family_role=data['peran_keluarga'],
                        birth_order=data.get('urutan_anak'),
                        phone_number=data.get('nomor_hp', ''),
                        email=data.get('email', ''),
                        blood_type=data.get('golongan_darah'),
                        baptism_date=tanggal_baptis,
                        sidi_date=tanggal_sidi,
                        marriage_date=tanggal_nikah,
                        membership_status=data.get('status_keanggotaan', 'FULL'),
                        is_deceased=data.get('sudah_meninggal', False),
                        deceased_date=tanggal_meninggal,
                        deceased_reason=data.get('sebab_meninggal', ''),
                    )
                    
                    success_count += 1
                    
                except Exception as e:
                    errors.append({
                        'row': row_num,
                        'error': str(e)
                    })
                    error_count += 1
            
            # If there are errors, rollback
            if error_count > 0:
                transaction.set_rollback(True)
                raise Exception('Ada error saat import')
        
        # Clear session
        request.session.pop('import_data', None)
        request.session.pop('import_errors', None)
        
        messages.success(
            request,
            f'Berhasil import {success_count} jemaat!'
        )
        return redirect('members:list')
        
    except Exception as e:
        context = {
            'step': 'error',
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors,
        }
        return render(request, 'members/bulk_import.html', context)


def parse_row(row, row_num):
    """
    Parse single row from Excel into dict.
    
    Row structure (0-indexed):
    0: nama_lengkap
    1: jenis_kelamin (M/F)
    2: tanggal_lahir (YYYY-MM-DD)
    3: nama_keluarga
    4: peran_keluarga (HUSBAND/WIFE/CHILD/OTHER)
    5: urutan_anak (number, optional)
    6: nama_sektor
    7: nomor_hp
    8: email
    9: golongan_darah
    10: tanggal_baptis
    11: tanggal_sidi
    12: tanggal_nikah
    13: status_keanggotaan
    14: alamat_jalan
    15: kota
    16: provinsi
    17: kode_pos
    18: telepon_keluarga
    19: sudah_meninggal (YA/TIDAK)
    20: tanggal_meninggal
    21: sebab_meninggal
    """
    
    errors = {}
    
    # Required fields
    if not row[0]:
        errors['nama_lengkap'] = 'Nama lengkap wajib diisi'
    if not row[1] or row[1] not in ['M', 'F']:
        errors['jenis_kelamin'] = 'Jenis kelamin harus M atau F'
    if not row[2]:
        errors['tanggal_lahir'] = 'Tanggal lahir wajib diisi'
    if not row[3]:
        errors['nama_keluarga'] = 'Nama keluarga wajib diisi'
    if not row[4] or row[4] not in ['HUSBAND', 'WIFE', 'CHILD', 'PARENT', 'OTHER']:
        errors['peran_keluarga'] = 'Peran keluarga tidak valid'
    if not row[6]:
        errors['nama_sektor'] = 'Nama sektor wajib diisi'
    if not row[14]:
        errors['alamat_jalan'] = 'Alamat jalan wajib diisi'
    if not row[15]:
        errors['kota'] = 'Kota wajib diisi'
    if not row[16]:
        errors['provinsi'] = 'Provinsi wajib diisi'
    if not row[18]:
        errors['telepon_keluarga'] = 'Telepon keluarga wajib diisi'
    
    if errors:
        raise ValidationError(errors)
    
    # Parse dates
    def parse_date(value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        try:
            return datetime.strptime(str(value), '%Y-%m-%d').date()
        except:
            return None
    
    # Parse deceased status
    is_deceased = False
    if row[19]:
        is_deceased = str(row[19]).upper() in ['YA', 'YES', 'TRUE', '1']
    
    # Validate: if CHILD, must have birth_order
    if row[4] == 'CHILD' and not row[5]:
        raise ValidationError({'urutan_anak': 'Anak wajib memiliki urutan kelahiran'})
    
    # Validate: if deceased, must have date
    if is_deceased and not row[20]:
        raise ValidationError({'tanggal_meninggal': 'Tanggal meninggal wajib diisi jika sudah meninggal'})
    
    return {
        'nama_lengkap': str(row[0]).strip(),
        'jenis_kelamin': row[1],
        'tanggal_lahir': parse_date(row[2]),
        'nama_keluarga': str(row[3]).strip(),
        'peran_keluarga': row[4],
        'urutan_anak': int(row[5]) if row[5] else None,
        'nama_sektor': str(row[6]).strip(),
        'nomor_hp': str(row[7]).strip() if row[7] else '',
        'email': str(row[8]).strip() if row[8] else '',
        'golongan_darah': row[9] if row[9] else None,
        'tanggal_baptis': parse_date(row[10]),
        'tanggal_sidi': parse_date(row[11]),
        'tanggal_nikah': parse_date(row[12]),
        'status_keanggotaan': row[13] if row[13] else 'FULL',
        'alamat_jalan': str(row[14]).strip(),
        'kota': str(row[15]).strip(),
        'provinsi': str(row[16]).strip(),
        'kode_pos': str(row[17]).strip() if row[17] else '',
        'telepon_keluarga': str(row[18]).strip(),
        'sudah_meninggal': is_deceased,
        'tanggal_meninggal': parse_date(row[20]) if is_deceased else None,
        'sebab_meninggal': str(row[21]).strip() if row[21] else '',
    }


@login_required
def download_template(request):
    """Download Excel template for bulk import"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Import Jemaat"
    
    # Headers
    headers = [
        'nama_lengkap', 'jenis_kelamin', 'tanggal_lahir', 'nama_keluarga',
        'peran_keluarga', 'urutan_anak', 'nama_sektor', 'nomor_hp',
        'email', 'golongan_darah', 'tanggal_baptis', 'tanggal_sidi',
        'tanggal_nikah', 'status_keanggotaan', 'alamat_jalan', 'kota',
        'provinsi', 'kode_pos', 'telepon_keluarga', 'sudah_meninggal',
        'tanggal_meninggal', 'sebab_meninggal'
    ]
    
    # Header row styling
    header_fill = PatternFill(start_color="6272f5", end_color="6272f5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Example data row
    example = [
        'Budi Santoso', 'M', '1980-05-15', 'Keluarga Santoso',
        'HUSBAND', '', 'Sektor 1', '08123456789',
        'budi@email.com', 'A+', '1990-06-10', '1995-08-20',
        '2005-12-25', 'FULL', 'Jl. Merdeka No. 123', 'Jakarta',
        'DKI Jakarta', '12345', '021-12345678', 'TIDAK',
        '', ''
    ]
    
    for col_num, value in enumerate(example, 1):
        ws.cell(row=2, column=col_num, value=value)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=template_import_jemaat.xlsx'
    
    wb.save(response)
    return response